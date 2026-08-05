"""
Excel import for bulk job entry, and a matching downloadable template.

The uploaded .xlsx is parsed entirely in memory (openpyxl reads directly
from the uploaded bytes) and discarded immediately after - only the
extracted row data gets written to Mongo, never the spreadsheet file
itself.

Expected columns (case-insensitive, order doesn't matter - matched by
header name), in the same order as the downloadable template:
    Company Name | Location | Job Description | HR Email | Role Name | Job URL | Salary
"Salary" is optional - every other column should be filled in per row.
"""
import io
from datetime import datetime, timezone
import openpyxl

from app.database.mongo import get_db
from app.services.job_matching import extract_email

# Maps expected column headers (lowercased, whitespace-stripped) to the job field they fill.
COLUMN_MAP = {
    "company name": "company",
    "location": "location",
    "job description": "description",
    "hr email": "hr_email",
    "role name": "title",
    "job url": "url",
    "salary": "salary_text",
}
REQUIRED_FIELDS = {"company", "description", "title"}  # url/location/hr_email/salary can be blank


def generate_sample_excel() -> bytes:
    """Builds the downloadable template in memory - never touches disk."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = ["Company Name", "Location", "Job Description", "HR Email", "Role Name", "Job URL", "Salary"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)

    ws.append(
        [
            "Acme Corp",
            "Chennai, India",
            "We are looking for a Senior Angular Developer with .NET Core and C# experience...",
            "hr@acmecorp.com",
            "Senior Angular Developer",
            "https://acmecorp.com/careers/123",
            "12,00,000 - 15,00,000",
        ]
    )
    ws.append(
        [
            "Example Pvt Ltd",
            "Bangalore, India",
            "Full Stack .NET Engineer role building Clean Architecture APIs...",
            "careers@example.com",
            "Full Stack .NET Engineer",
            "https://example.com/jobs/456",
            "",  # salary is optional - fine to leave blank
        ]
    )

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(18, len(header) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(value) -> str:
    return str(value or "").strip().lower()


async def import_excel_jobs(file_content: bytes) -> dict:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    header_row = rows[0]
    col_index_to_field = {}
    for idx, header in enumerate(header_row):
        field = COLUMN_MAP.get(_normalize_header(header))
        if field:
            col_index_to_field[idx] = field

    missing_required = REQUIRED_FIELDS - set(col_index_to_field.values())
    if missing_required:
        raise ValueError(
            f"Missing required column(s): {', '.join(missing_required)}. "
            f"Download the sample template to see the expected format."
        )

    db = get_db()
    imported = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue  # skip fully blank rows

        job = {"source": "excel_import", "salary_text": ""}
        for idx, field in col_index_to_field.items():
            value = row[idx] if idx < len(row) else None
            job[field] = str(value).strip() if value is not None else ""

        missing = [f for f in REQUIRED_FIELDS if not job.get(f)]
        if missing:
            skipped += 1
            errors.append(f"Row {row_num}: missing {', '.join(missing)} - skipped")
            continue

        if not job.get("hr_email"):
            job["hr_email"] = extract_email(job.get("description", "")) or ""

        now = datetime.now(timezone.utc)
        await db.jobs.insert_one(
            {
                **job,
                "status": "new",
                "notes": "",
                "analysis": None,
                "hr_email_guess": job.get("hr_email") or extract_email(job.get("description", "")),
                "created_at": now,
                "updated_at": now,
            }
        )
        imported += 1

    return {"imported": imported, "skipped": skipped, "errors": errors}
